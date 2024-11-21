def process_swap_request(request_data):
    # Validate the request
    if not validate_date_range(request_data['current_shift'], request_data['desired_shift']):
        return "Invalid date range"

    if not check_shift_availability(request_data['current_shift'], request_data['desired_shift']):
        return "Shifts not available"

    # Check for overlaps
    if check_shift_overlap(request_data['current_shift'], request_data['desired_shift']):
        return "Shift overlap detected"

    # Notify employees
    send_notification(employee1_email, employee2_email, request_data)

    # Update the schedule (if approved)
    update_schedule(request_data)

    return "Swap request processed successfully"

from flask import Flask, render_template

app = Flask(__name__)

@app.route('/')
def index():
    data = {'name': 'Alice', 'age': 30}
    return render_template('index.html', data=data)

from flask import Flask, request, jsonify

app = Flask(__name__)

# Assuming you have a function to retrieve an employee's schedule
def get_employee_schedule(employee_id):
    # ... (logic to fetch schedule from database or API)
    return schedule

@app.route('/swap_request', methods=['POST'])
def process_swap_request():
    employee_id = request.json['employee_id']
    new_shift_start = request.json['new_shift_start']
    new_shift_end = request.json['new_shift_end']

    schedule = get_employee_schedule(employee_id)

    for shift in schedule:
        if (new_shift_start <= shift['end_time'] and new_shift_end >= shift['start_time']):
            return jsonify({'error': 'Overlap detected'}), 400

    # If no overlap, proceed with the swap (e.g., update database, send notifications)
    return jsonify({'message': 'Swap request processed successfully'}), 200


    import openpyxl

def fill_excel_sheet(employee_id, sutherland_email, current_shift, desired_shift):
    # Load the existing Excel workbook
    workbook = openpyxl.load_workbook('your_excel_file.xlsx')
    sheet = workbook.active  # Assuming the active sheet is the target

    # Find the next empty row
    next_row = sheet.max_row + 1

    # Write the data to the next row
    sheet.cell(row=next_row, column=1).value = employee_id
    sheet.cell(row=next_row, column=2).value = sutherland_email
    sheet.cell(row=next_row, column=3).value = current_shift
    sheet.cell(row=next_row, column=4).value = desired_shift

    # Save the workbook
    workbook.save('your_excel_file.xlsx')

# Example usage (replace with actual form data)
employee_id = "EMP123"
sutherland_email = "emp123@sutherland.com"
current_shift = "08:00-16:00"
desired_shift = "16:00-24:00"

fill_excel_sheet(employee_id, sutherland_email, current_shift, desired_shift)

from flask import Flask, request, render_template

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST': 1 
        employee_id = request.form['employee_id']
        sutherland_email = request.form['sutherland_email']
        current_shift = request.form['current_shift']
        desired_shift = request.form['desired_shift']

        fill_excel_sheet(employee_id, sutherland_email, current_shift, desired_shift)
        return "Data added to Excel successfully!"

    return render_template('form.html')

if __name__ == '__main__':
    app.run(debug=True)

   